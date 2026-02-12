import json
import boto3
from datetime import datetime, timedelta
import base64
import io

import openpyxl
from fpdf import FPDF

# DynamoDB
dynamodb = boto3.resource("dynamodb")
attendance_table = dynamodb.Table("Attendance")
users_table = dynamodb.Table("Users")


def response(status, body, content_type="application/json", is_binary=False, filename=None):
    headers = {
        "Access-Control-Allow-Origin": "*",
        "Access-Control-Allow-Headers": "*",
        "Access-Control-Allow-Methods": "GET,OPTIONS",
        "Content-Type": content_type
    }
    if filename:
        headers["Content-Disposition"] = f'attachment; filename="{filename}"'

    return {
        "statusCode": status,
        "headers": headers,
        "isBase64Encoded": is_binary,
        "body": body
    }


def lambda_handler(event, context):

    # ---------- CORS ----------
    if event.get("httpMethod") == "OPTIONS":
        return response(200, json.dumps({}))

    params = event.get("queryStringParameters") or {}

    from_date = params.get("fromDate")
    to_date = params.get("toDate")
    role_filter = params.get("role") or "All"
    status_filter = params.get("status") or "All"
    output_format = params.get("format")

    if not from_date or not to_date:
        return response(400, json.dumps({"message": "fromDate and toDate required"}))

    start = datetime.strptime(from_date, "%Y-%m-%d")
    end = datetime.strptime(to_date, "%Y-%m-%d")

    users = users_table.scan().get("Items", [])
    attendance_items = attendance_table.scan().get("Items", [])

    records = []
    present_count = 0
    absent_count = 0

    current = start
    while current <= end:
        date_str = current.strftime("%d-%m-%Y")

        for user in users:
            if role_filter != "All" and user["role"] != role_filter:
                continue

            present = any(
                a["email"] == user["email"] and a["date"] == current.strftime("%Y-%m-%d")
                for a in attendance_items
            )

            status = "Present" if present else "Absent"

            if status_filter != "All" and status != status_filter:
                continue

            if present:
                present_count += 1
            else:
                absent_count += 1

            records.append({
                "date": date_str,
                "name": user["name"],
                "role": user["role"],
                "registerNumber": user.get("registerNumber") or "-",
                "status": status
            })

        current += timedelta(days=1)

    # ===================== CSV =====================
    if output_format == "csv":
        csv = "Date,Name,Role,Register Number,Status\n"
        for r in records:
            csv += f"{r['date']},{r['name']},{r['role']},{r['registerNumber']},{r['status']}\n"

        return response(
            200,
            csv,
            "text/csv",
            False,
            "attendance.csv"
        )

    # ===================== EXCEL =====================
    if output_format == "excel":
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Attendance"

        ws.append(["Date", "Name", "Role", "Register Number", "Status"])

        for r in records:
            ws.append([
                r["date"],
                r["name"],
                r["role"],
                r["registerNumber"],
                r["status"]
            ])

        stream = io.BytesIO()
        wb.save(stream)
        stream.seek(0)

        encoded = base64.b64encode(stream.read()).decode()

        return response(
            200,
            encoded,
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            True,
            "attendance.xlsx"
        )

    # ===================== PDF =====================
    if output_format == "pdf":
        pdf = FPDF()
        pdf.set_auto_page_break(auto=True, margin=15)
        pdf.add_page()

        # Header
        pdf.set_font("Arial", "B", 16)
        pdf.cell(0, 10, "SMART ATTENDANCE SYSTEM", ln=True, align="C")
        pdf.set_font("Arial", "B", 12)
        pdf.cell(0, 8, "Attendance Report", ln=True, align="C")
        pdf.ln(6)

        # Report details
        pdf.set_font("Arial", size=10)
        pdf.cell(0, 6, f"Report Generated On : {datetime.now().strftime('%d-%m-%Y')}", ln=True)
        pdf.cell(0, 6, f"Report Period      : {from_date} to {to_date}", ln=True)
        pdf.cell(0, 6, f"Role Filter        : {role_filter}", ln=True)
        pdf.cell(0, 6, f"Status Filter      : {status_filter}", ln=True)
        pdf.ln(6)

        # Summary
        pdf.set_font("Arial", "B", 11)
        pdf.cell(0, 6, "Summary", ln=True)
        pdf.set_font("Arial", size=10)
        pdf.cell(0, 6, f"Total Records : {len(records)}", ln=True)
        pdf.cell(0, 6, f"Present       : {present_count}", ln=True)
        pdf.cell(0, 6, f"Absent        : {absent_count}", ln=True)
        pdf.ln(6)

        # Table header
        pdf.set_font("Arial", "B", 9)
        col_widths = [25, 40, 25, 35, 25]
        headers = ["Date", "Name", "Role", "Register No", "Status"]

        for i in range(len(headers)):
            pdf.cell(col_widths[i], 8, headers[i], border=1, align="C")
        pdf.ln()

        # Table rows
        pdf.set_font("Arial", size=9)
        for r in records:
            pdf.cell(col_widths[0], 8, r["date"], border=1)
            pdf.cell(col_widths[1], 8, r["name"], border=1)
            pdf.cell(col_widths[2], 8, r["role"], border=1)
            pdf.cell(col_widths[3], 8, r["registerNumber"], border=1)
            pdf.cell(col_widths[4], 8, r["status"], border=1)
            pdf.ln()

        pdf.ln(8)
        pdf.set_font("Arial", size=9)
        pdf.cell(0, 6, "Generated by Smart Attendance System", align="C")

        pdf_bytes = pdf.output(dest="S").encode("latin-1")
        encoded = base64.b64encode(pdf_bytes).decode()

        return response(
            200,
            encoded,
            "application/pdf",
            True,
            "attendance.pdf"
        )

    # ===================== JSON PREVIEW =====================
    return response(200, json.dumps(records))
